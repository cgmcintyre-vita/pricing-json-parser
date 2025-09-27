from pricingreport import AMERPricingFile, ItemPrice
from pathlib import Path
import openpyxl as pyxl
from SAPCCReport import SAPCCReportFile
from openpyxl.worksheet.worksheet import Worksheet


def print_header(output_sheet:Worksheet):
    header_items = ["SKU", "US Price", "US Sale Price", "CA Price", "CA Sale Price"]
    for col_num, header in enumerate(header_items, start=1):
        output_sheet.cell(row=1, column=col_num, value=header)
def print_item_price(output_sheet:Worksheet, ip:ItemPrice):
    output_sheet.append(ip.output_str_value)

def print_prices(input_path:Path, output_path:Path):
    print("Printing prices")
    input_file = SAPCCReportFile(input_path, pyxl.load_workbook(input_path))
    
    pricing_report = AMERPricingFile(input_file)
    prices = pricing_report.get_prices()
    
    #print output
    output_wb = pyxl.Workbook()
    ws = output_wb.active
    if isinstance(ws, Worksheet):
        print("printing headers")
        print_header(ws)

        for price in prices:
            print_item_price(ws, price)
    else:
        raise TypeError("Received wrong type of workbook")
    
    output_wb.save(output_path)



    